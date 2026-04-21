from django.db import connection, transaction
from django.views.decorators.http import require_GET, require_POST
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from openpyxl import load_workbook
import io
import re
from authentification.decoratos import admin_required
from commercial.metier.Category import Category
from commercial.metier.Product import Product
from commercial.metier.ProductCategory import ProductCategory
from commercial.metier.Unit import Unit
from unidecode import unidecode

@require_GET
@admin_required
def import_page(request):
    return render(
        request, 
        "views/import.html"
    )

@require_GET
@admin_required
def categories_api(request):
    categories = Category.objects.order_by('name').values('id', 'name')
    return JsonResponse({'categories': list(categories)})

@require_POST
@admin_required
def read_excel_file(request):
    excel_file = request.FILES.get('excel_file')
    category_mode = request.POST.get('category_mode', 'existing').strip().lower()
    product_mode = request.POST.get('product_mode', 'new').strip().lower()
    existing_category_id = request.POST.get('existing_category_id', '').strip()
    new_category_name = request.POST.get('new_category_name', '').strip()
    
    if not excel_file:
        messages.error(request, "Aucun fichier sélectionné")
        return redirect('import_page')
    
    if not excel_file.name.lower().endswith(('.xlsx', '.xlsm')):
        messages.error(request, "Format non supporté. Utilisez un fichier .xlsx ou .xlsm")
        return redirect('import_page')
    
    try:
        workbook = load_workbook(filename=io.BytesIO(excel_file.read()), data_only=True)
        sheet = workbook.active

        if category_mode not in {'existing', 'new'}:
            messages.error(request, "Mode de catégorie invalide.")
            return redirect('import_page')

        if product_mode not in {'new', 'existing'}:
            product_mode = 'new'

        selected_category = None
        if category_mode == 'existing':
            if not existing_category_id:
                messages.error(request, "Veuillez sélectionner une catégorie existante.")
                return redirect('import_page')

            selected_category = Category.objects.filter(id=existing_category_id).first()
            if selected_category is None:
                messages.error(request, "La catégorie sélectionnée est introuvable.")
                return redirect('import_page')
        
        else:
            if new_category_name:
                category_names = [new_category_name]
            else:
                raw_category = sheet.cell(row=2, column=9).value
                if not isinstance(raw_category, str) or not raw_category.strip():
                    messages.error(request, "La catégorie est introuvable (cellule I2 vide ou invalide).")
                    return redirect('import_page')

                category_name_source = ' '.join(raw_category.strip().split()[:-1]).strip()
                if not category_name_source:
                    category_name_source = raw_category.strip()

                category_names = [
                    part.strip()
                    for part in re.split(r'[;,/|]+', category_name_source)
                    if part.strip()
                ]
                if not category_names:
                    category_names = [category_name_source]

        if category_mode == 'existing':
            category_names = [selected_category.name]

        nb_vide = 0
        products_to_insert = []  # Liste pour stocker toutes les données
        row_errors = []
        invalid_rows = set()
        
        for excel_row in range(7, sheet.max_row + 1):
            has_data = False
            
            # Vérifier si la ligne a des données
            for col_idx in range(1, 13):
                cell = sheet.cell(row=excel_row, column=col_idx)
                cell_value = cell.value
                
                if isinstance(cell_value, str):
                    cell_value = cell_value.strip()
                
                if cell_value is not None and cell_value != '' and str(cell_value).strip() != '' and cell_value != 0:
                    has_data = True
            
            if has_data:
                nb_vide = 0
                try:
                    designation_raw = sheet.cell(row=excel_row, column=2).value
                    unite_raw = sheet.cell(row=excel_row, column=7).value
                    quantite = sheet.cell(row=excel_row, column=9).value
                    prixAchat = sheet.cell(row=excel_row, column=10).value
                    coefficient = sheet.cell(row=excel_row, column=11).value
                    prixVente = sheet.cell(row=excel_row, column=12).value

                    if not isinstance(designation_raw, str) or not designation_raw.strip():
                        raise ValueError("désignation vide")
                    if not isinstance(unite_raw, str) or not unite_raw.strip():
                        raise ValueError("unité vide")
                    if prixAchat is None:
                        raise ValueError("prix d'achat manquant")
                    if coefficient is None:
                        raise ValueError("coefficient manquant")
                    if prixVente is None:
                        raise ValueError("prix de vente manquant")

                    designation = unidecode(designation_raw.strip()).lower()
                    unite = unidecode(unite_raw.strip()).lower()

                    try:
                        if quantite is not None and str(quantite).strip() != '':
                            quantite = float(quantite)
                    except (TypeError, ValueError):
                        raise ValueError("quantité invalide")

                    try:
                        prixAchat = float(prixAchat)
                    except (TypeError, ValueError):
                        raise ValueError("prix d'achat invalide")

                    try:
                        coefficient = float(coefficient)
                    except (TypeError, ValueError):
                        raise ValueError("coefficient invalide")

                    try:
                        prixVente = float(prixVente)
                    except (TypeError, ValueError):
                        raise ValueError("prix de vente invalide")

                    products_to_insert.append({
                        'excel_row': excel_row,
                        'designation': designation,
                        'unite': unite,
                        'quantite': quantite,
                        'prixAchat': prixAchat,
                        'coefficient': coefficient,
                        'prixVente': prixVente,
                    })
                except Exception as row_error:
                    row_errors.append(f"Ligne Excel {excel_row}: {row_error}")
                    invalid_rows.add(excel_row)
            else:
                nb_vide += 1
            
            if nb_vide >= 2:
                break

        seen_designations = {}
        for product in products_to_insert:
            designation = product['designation']
            excel_row = product['excel_row']

            if designation in seen_designations:
                first_row = seen_designations[designation]
                row_errors.append(
                    f"Ligne Excel {excel_row}: désignation dupliquée dans le fichier "
                    f"(déjà présente à la ligne {first_row}) -> '{designation}'."
                )
                invalid_rows.add(excel_row)
            else:
                seen_designations[designation] = excel_row

        candidate_products = [
            product for product in products_to_insert
            if product['excel_row'] not in invalid_rows
        ]

        products_to_insert = [
            product for product in candidate_products
            if product['excel_row'] not in invalid_rows
        ]

        if row_errors:
            messages.error(
                request,
                f"Import annulé: {len(row_errors)} erreur(s) détectée(s). Corrigez-les puis réimportez."
            )
            for row_error in row_errors:
                messages.error(request, row_error, extra_tags='excel-row-error')
            return redirect('import_page')
        
        if products_to_insert:
            with transaction.atomic():
                if category_mode == 'existing':
                    categories = [selected_category]
                else:
                    categories = [Category.objects.get_or_create(name=category_name)[0] for category_name in category_names]

                unit_names = sorted({product['unite'] for product in products_to_insert if product['unite']})
                units = {
                    unit.name: unit
                    for unit in Unit.objects.filter(name__in=unit_names)
                }

                for unit_name in unit_names:
                    if unit_name not in units:
                        units[unit_name] = Unit.objects.create(name=unit_name)

                existing_products = {
                    product.designation: product
                    for product in Product.objects.filter(
                        designation__in=[product['designation'] for product in products_to_insert]
                    ).only('id', 'designation')
                }

                product_objects = []
                product_rows = []
                skipped_missing_existing_products = 0
                for product_data in products_to_insert:
                    existing_product = existing_products.get(product_data['designation'])
                    if existing_product is not None:
                        if product_mode == 'existing':
                            existing_product.purchase_unit_price = product_data['prixAchat']
                            existing_product.sale_unit_price = product_data['prixVente']
                            existing_product.coefficient = product_data['coefficient']
                            unit_obj = units.get(product_data['unite'])
                            if unit_obj is not None and existing_product.unit_id != unit_obj.id:
                                existing_product.unit = unit_obj
                            existing_product.save(update_fields=['purchase_unit_price', 'sale_unit_price', 'coefficient', 'unit'])
                            product_rows.append(existing_product)
                            continue

                        product_rows.append(existing_product)
                        continue

                    if product_mode == 'existing':
                        skipped_missing_existing_products += 1
                        continue

                    unit_obj = units.get(product_data['unite'])
                    if unit_obj is None:
                        continue

                    product_objects.append(Product(
                        designation=product_data['designation'],
                        purchase_unit_price=product_data['prixAchat'],
                        sale_unit_price=product_data['prixVente'],
                        coefficient=product_data['coefficient'],
                        unit=unit_obj,
                    ))
                    product_rows.append(product_data)

                created_products = Product.objects.bulk_create(product_objects)
                new_products_by_designation = {
                    product.designation: product
                    for product in created_products
                }
                product_category_objects = []
                linked_existing_products = 0
                updated_existing_products = 0
                for product_row in products_to_insert:
                    product = existing_products.get(product_row['designation']) or new_products_by_designation.get(product_row['designation'])
                    if product is None:
                        continue

                    if product_mode == 'existing' and product_row['designation'] in existing_products:
                        updated_existing_products += 1

                    for category in categories:
                        product_category_objects.append(
                            ProductCategory(product=product, category=category)
                        )

                    if product_row['designation'] in existing_products:
                        linked_existing_products += 1

                if product_category_objects:
                    ProductCategory.objects.bulk_create(product_category_objects, ignore_conflicts=True)

                count = len(created_products)
                category_label = ', '.join(category.name for category in categories)
                if product_mode == 'existing':
                    if count and updated_existing_products:
                        messages.success(
                            request,
                            f"Import réussi - {count} produit(s) créés et {updated_existing_products} produit(s) existant(s) mis à jour et reliés aux catégories {category_label}"
                        )
                    elif updated_existing_products:
                        messages.success(
                            request,
                            f"Import réussi - {updated_existing_products} produit(s) existant(s) mis à jour et reliés aux catégories {category_label}"
                        )
                    elif skipped_missing_existing_products:
                        messages.warning(
                            request,
                            f"Aucun produit correspondant trouvé pour {skipped_missing_existing_products} désignation(s) du fichier."
                        )
                    else:
                        messages.warning(request, "Aucun produit existant n'a pu être mis à jour.")
                else:
                    if count and linked_existing_products:
                        messages.success(
                            request,
                            f"Import réussi - {count} produit(s) créés et {linked_existing_products} produit(s) existant(s) relié(s) aux catégories {category_label}"
                        )
                    elif count:
                        messages.success(request, f"Import réussi - {count} produits ajoutés dans la catégorie {category_label}")
                    elif linked_existing_products:
                        messages.success(
                            request,
                            f"Import réussi - {linked_existing_products} produit(s) existant(s) relié(s) aux catégories {category_label}"
                        )
                    else:
                        messages.warning(request, "Aucun produit nouveau ni liaison de catégorie n'a pu être créé.")
        else:
            messages.warning(request, "Aucune donnée valide à importer")
            return redirect('import_page')
        
    except Exception as e:
        print(f"Erreur: {e}")
        error_message = str(e)

        if "product_designation_key" in error_message or "(designation)=" in error_message:
            duplicate_match = re.search(r"\(designation\)=\((.*?)\)", error_message)
            duplicate_designation = duplicate_match.group(1) if duplicate_match else None

            if duplicate_designation:
                messages.error(
                    request,
                    f"Erreur d'import: la désignation '{duplicate_designation}' existe déjà. "
                    f"Supprimez-la du fichier ou modifiez-la avant de réimporter."
                )
            else:
                messages.error(
                    request,
                    "Erreur d'import: une désignation existe déjà en base. "
                    "Corrigez les doublons puis réessayez."
                )
        else:
            messages.error(request, f"Erreur de lecture: {error_message}")
        return redirect('import_page')
    
    return redirect('liste_product_page')